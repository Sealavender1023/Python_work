"""
说明： 
本工具专为高分影像生产场景设计，核心功能是高效解决分幅影像元数据的批量生成难题。
它支持从指定的边界数据和目标数据文件中提取必要信息，并生成符合要求的元数据文件。
支持的输入格式包括.xls和.xlsx，输出格式同样支持这两种格式。
功能特点：
1. 支持拖放操作，简化用户交互。
2. 自动校验输入文件格式，确保数据一致性。
3. 生成详细的运行日志，便于追踪处理过程。


Copyright (c) 2025/07/25 Xiongxiao Wu.

"""
import os
import re
import shutil
import pandas as pd
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import xlwt
from tkinterdnd2 import DND_FILES, TkinterDnD
from typing import Optional, Dict, List, Set, Tuple
import logging
from pathlib import Path
import numpy as np
import xlrd  # 用于读取.xls文件
import openpyxl  # 用于读取.xlsx文件
from openpyxl.utils import get_column_letter
from xlutils.copy import copy


class MetadataProcessor:
    """元数据处理器，封装核心处理逻辑"""

    def __init__(self, template_path: str):
        self.template_path = template_path
        self.boundary_data: Dict[str, List] = None  # 原始范围数据
        self.metadata_data: Dict[str, List] = None  # 目标区域数据
        self.output_format = '.xlsx'  # 默认输出格式

    def load_data(self, boundary_file: str, target_file: str) -> None:
        """加载边界数据和目标数据"""
        # 根据目标文件确定输出格式
        self.output_format = Path(target_file).suffix.lower()
        if self.output_format not in ['.xls', '.xlsx']:
            self.output_format = '.xlsx'  # 默认使用xlsx格式

        self.boundary_data = self._process_excel(boundary_file)
        self.metadata_data = self._process_excel(target_file)

    @staticmethod
    def _process_excel(file_path: str) -> Dict[str, List]:
        """处理Excel文件核心方法，支持.xls和.xlsx格式"""
        try:
            # 根据文件扩展名选择读取方式
            if file_path.lower().endswith('.xls'):
                # 使用xlrd读取.xls文件
                workbook = xlrd.open_workbook(file_path)
                sheet = workbook.sheet_by_index(0)

                # 获取列索引
                headers = sheet.row_values(0)
                map_index_idx = headers.index('MapIndex') if 'MapIndex' in headers else -1
                time_idx = headers.index('time') if 'time' in headers else -1

                if map_index_idx == -1 or time_idx == -1:
                    raise ValueError("Excel文件缺少必要列: MapIndex 或 time")

                map_index = []
                time_vals = []
                for row_idx in range(1, sheet.nrows):
                    row = sheet.row_values(row_idx)
                    map_index.append(str(row[map_index_idx]))
                    time_vals.append(row[time_idx])

                data = {'MapIndex': map_index, 'time': time_vals}
                df = pd.DataFrame(data)
            else:
                # 使用pandas读取.xlsx文件
                df = pd.read_excel(file_path)
                if 'MapIndex' not in df or 'time' not in df:
                    raise ValueError("Excel文件缺少必要列: MapIndex 或 time")

            return MetadataProcessor._extract_metadata(df)
        except FileNotFoundError as e:
            raise RuntimeError(f"文件不存在: {file_path}") from e
        except Exception as e:
            raise RuntimeError(f"处理文件出错: {str(e)}") from e

    @staticmethod
    def _extract_metadata(df: pd.DataFrame) -> Dict[str, List]:
        """从DataFrame中提取元数据"""
        map_index = df['MapIndex'].astype(str).tolist()
        return {
            'file_name': map_index,
            'row_number': [s[:4] for s in map_index],
            'column_number': [s[7:12] for s in map_index],
            'band_number': [f"{int(s[7:9])}°" for s in map_index],
            'central_meridian': [f"{int(value) * 3}°" for value in [s[7:9] for s in map_index]],
            'flight_times': df['time'].tolist()
        }

    def generate_metadata(self, tif_folder: str, output_dir: str) -> None:
        """生成元数据文件主方法"""
        self._add_tif_sizes(tif_folder)
        self._generate_coordinates()
        self._process_boundary_connections()
        self._generate_output(output_dir)

    def _add_tif_sizes(self, tif_folder: str) -> None:
        """添加TIFF文件大小信息"""
        sizes = []
        for name in self.metadata_data['file_name']:
            tif_path = os.path.join(tif_folder, f"{name}.tif")
            if os.path.isfile(tif_path):
                size_mb = os.path.getsize(tif_path) / (1024 ** 2)
                sizes.append(f"{size_mb:.2f}MB")
            else:
                sizes.append("数据不存在")
                print(f"警告: 文件 {name}.tif 不存在")
        self.metadata_data['tif_size'] = sizes

    def _generate_coordinates(self) -> None:
        """生成图幅角点坐标"""
        coord_data = self.metadata_data
        row = [int(r) for r in coord_data['row_number']]
        col = [int(c) for c in coord_data['column_number']]

        coord_data.update({
            'WS_X': [f"{r * 1000}.00" for r in row],
            'WS_Y': [f"{c * 1000}.00" for c in col],
            'WN_X': [f"{(r + 1) * 1000}.00" for r in row],
            'WN_Y': [f"{c * 1000}.00" for c in col],
            'EN_X': [f"{(r + 1) * 1000}.00" for r in row],
            'EN_Y': [f"{(c + 1) * 1000}.00" for c in col],
            'ES_X': [f"{r * 1000}.00" for r in row],
            'ES_Y': [f"{(c + 1) * 1000}.00" for c in col],
            'filename': [f"文件:{name}" for name in coord_data['file_name']]
        })

    def _process_boundary_connections(self) -> None:
        """处理图幅接边关系"""
        boundary_coord = self._get_boundary_coordinates()
        self._process_directional_connections(boundary_coord)
        self._process_diagonal_connections(boundary_coord)

    def _get_boundary_coordinates(self) -> Set[Tuple[int, int]]:
        """获取边界坐标集合"""
        return set(zip(
            [int(r) for r in self.boundary_data['row_number']],
            [int(c) for c in self.boundary_data['column_number']]
        ))

    def _process_directional_connections(self, boundary: Set[Tuple[int, int]]) -> None:
        """处理四个方向的接边关系"""
        directions = {
            'N': (1, 0),
            'S': (-1, 0),
            'E': (0, 1),
            'W': (0, -1)
        }

        for direction, (dr, dc) in directions.items():
            target_coord = [
                (int(r) + dr, int(c) + dc)
                for r, c in zip(
                    self.metadata_data['row_number'],
                    self.metadata_data['column_number']
                )
            ]

            self.metadata_data[f'Link_{direction}'] = [
                "已接" if coord in boundary else "自由边"
                for coord in target_coord
            ]

            self.metadata_data[f'filename_{direction}'] = [
                self._generate_filename(r, c) if coord in boundary else "无"
                for coord, (r, c) in zip(target_coord, target_coord)
            ]

    def _process_diagonal_connections(self, boundary: Set[Tuple[int, int]]) -> None:
        """处理对角线方向的接边关系"""
        diagonals = {
            'WN': (1, -1),
            'EN': (1, 1),
            'WS': (-1, -1),
            'ES': (-1, 1)
        }

        for code, (dr, dc) in diagonals.items():
            target_coord = [
                (int(r) + dr, int(c) + dc)
                for r, c in zip(
                    self.metadata_data['row_number'],
                    self.metadata_data['column_number']
                )
            ]

            self.metadata_data[f'filename_{code}'] = [
                self._generate_filename(r, c) if coord in boundary else "无"
                for coord, (r, c) in zip(target_coord, target_coord)
            ]

    @staticmethod
    def _generate_filename(r: int, c: int) -> str:
        """生成标准文件名"""
        return f"{r}.0-{c}.0"

    def _generate_output(self, output_dir: str) -> None:
        """生成输出文件，格式与输入文件相同"""
        mapping = {
            'filename': 'A1',
            'file_name': 'C8',
            'tif_size': 'C17',
            'WS_X': 'C20', 'WS_Y': 'C21',
            'WN_X': 'C22', 'WN_Y': 'C23',
            'EN_X': 'C24', 'EN_Y': 'C25',
            'ES_X': 'C26', 'ES_Y': 'C27',
            'Link_W': 'C41', 'Link_N': 'C42',
            'Link_E': 'C43', 'Link_S': 'C44',
            'filename_WN': 'C45', 'filename_N': 'C46',
            'filename_EN': 'C47', 'filename_W': 'C48',
            'filename_E': 'C49', 'filename_WS': 'C50',
            'filename_S': 'C51', 'filename_ES': 'C52',
            'band_number': 'C37', 'central_meridian': 'C35',
            'flight_times': 'C68'
        }

        self._validate_data_consistency()
        os.makedirs(output_dir, exist_ok=True)

        for idx, orig_name in enumerate(self.metadata_data['file_name']):
            self._generate_single_file(idx, orig_name, mapping, output_dir)

    def _validate_data_consistency(self) -> None:
        """验证数据一致性"""
        lengths = {k: len(v) for k, v in self.metadata_data.items()}
        if len(set(lengths.values())) > 1:
            raise ValueError("数据字段长度不一致，请检查输入数据")

        if len(self.metadata_data['file_name']) != len(set(self.metadata_data['file_name'])):
            raise ValueError("清洗后的文件名存在重复")

    def _generate_single_file(self, index: int, orig_name: str,
                              mapping: Dict, output_dir: str) -> None:
        """生成单个元数据文件，格式与输入相同"""
        safe_name = re.sub(r'[\\/*?:"<>|]', '_', orig_name).strip()
        output_path = os.path.join(output_dir, f"{safe_name}{self.output_format}")

        try:
            # 复制模板文件
            shutil.copy(self.template_path, output_path)

            # 根据输出格式选择处理方式
            if self.output_format == '.xlsx':
                self._write_xlsx(output_path, index, mapping)
            elif self.output_format == '.xls':
                self._write_xls(output_path, index, mapping)
            else:
                raise ValueError(f"不支持的输出格式: {self.output_format}")

        except Exception as e:
            print(f"生成文件 {safe_name} 失败: {str(e)}")
            raise

    def _write_xlsx(self, output_path: str, index: int, mapping: Dict, orig_name=None):
        """写入.xlsx格式文件"""
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active

        current_data = {
            **{k: v[index] for k, v in self.metadata_data.items()},
            'original_filename': orig_name
        }

        for field, cell in mapping.items():
            if field in current_data:
                ws[cell] = current_data[field]

        wb.save(output_path)
        wb.close()

    def _write_xls(self, output_path: str, index: int, mapping: Dict, orig_name=None):
        """写入.xls格式文件"""
        # 打开复制的模板文件
        rb = xlrd.open_workbook(output_path, formatting_info=True)
        wb = copy(rb)  # 使用xlutils复制工作簿
        ws = wb.get_sheet(0)  # 获取第一个工作表

        current_data = {
            **{k: v[index] for k, v in self.metadata_data.items()},
            'original_filename': orig_name
        }

        # 创建样式对象保留原有格式
        style = xlwt.XFStyle()
        style.font = xlwt.Font()
        style.alignment = xlwt.Alignment()

        for field, cell in mapping.items():
            if field in current_data:
                # 转换单元格地址为行号和列号
                row, col = self._cell_address_to_indices(cell)
                ws.write(row, col, current_data[field], style)

        wb.save(output_path)

    @staticmethod
    def _cell_address_to_indices(cell_address: str) -> Tuple[int, int]:
        """将Excel单元格地址(如'A1')转换为(行索引, 列索引)"""
        # 分离列字母和行数字
        col_letter = ''.join(filter(str.isalpha, cell_address))
        row_num = int(''.join(filter(str.isdigit, cell_address)))

        # 将列字母转换为列索引 (0-based)
        col_index = 0
        for char in col_letter:
            col_index = col_index * 26 + (ord(char.upper()) - ord('A') + 1)
        col_index -= 1  # 转换为0-based索引

        # 行号转换为0-based索引
        row_index = row_num - 1

        return row_index, col_index


class MetadataProcessorGUI(TkinterDnD.Tk):
    """元数据处理器图形界面"""

    def __init__(self):
        super().__init__()
        self.title("元数据生成工具 v1.0")
        self.geometry("800x500")  # 调整窗口大小
        self.processor: Optional[MetadataProcessor] = None
        self.template_path = ""

        # 配置日志
        self.logger = logging.getLogger('APP')
        self._setup_ui()
        self._setup_logging()

    def _setup_ui(self):
        """初始化界面组件"""
        main_frame = ttk.Frame(self)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # 配置网格列权重
        main_frame.columnconfigure(0, weight=3)  # 左侧区域占3份
        main_frame.columnconfigure(1, weight=5)  # 右侧说明区域占2份
        main_frame.rowconfigure(0, weight=5)

        # 左侧输入区域
        left_frame = ttk.Frame(main_frame)
        left_frame.grid(row=0, column=0, sticky="nsew", padx=5)

        # 右侧说明区域
        right_frame = ttk.Frame(main_frame)
        right_frame.grid(row=0, column=1, sticky="nsew", padx=5)

        # ================= 左侧组件 =================
        ttk.Label(left_frame, text="模板文件:").grid(row=0, column=0, sticky=tk.W)
        self.template_entry = self._create_file_entry(left_frame, row=0, command=self.select_template)

        ttk.Label(left_frame, text="边界数据文件:").grid(row=1, column=0, sticky=tk.W)
        self.boundary_entry = self._create_file_entry(left_frame, row=1)

        ttk.Label(left_frame, text="目标数据文件:").grid(row=2, column=0, sticky=tk.W)
        self.target_entry = self._create_file_entry(left_frame, row=2)

        ttk.Label(left_frame, text="TIFF文件夹:").grid(row=3, column=0, sticky=tk.W)
        self.tiff_entry = self._create_file_entry(left_frame, row=3, is_folder=True)

        ttk.Label(left_frame, text="输出目录:").grid(row=4, column=0, sticky=tk.W)
        self.output_entry = self._create_file_entry(left_frame, row=4, is_folder=True)

        # 控制按钮
        btn_frame = ttk.Frame(left_frame)
        btn_frame.grid(row=5, column=0, columnspan=2, pady=10)
        ttk.Button(btn_frame, text="开始处理", command=self.start_processing).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="重置所有", command=self.reset_fields).pack(side=tk.LEFT, padx=5)

        # 日志区域
        log_frame = ttk.LabelFrame(left_frame, text="运行日志", padding=5)
        log_frame.grid(row=6, column=0, columnspan=2, sticky=tk.NSEW)
        self.log_text = tk.Text(log_frame, wrap=tk.WORD, state=tk.DISABLED)
        self.log_text.pack(fill=tk.BOTH, expand=True)

        # 配置左侧框架的行列权重
        left_frame.rowconfigure(6, weight=1)
        left_frame.columnconfigure(0, weight=1)

        # ================= 右侧说明区域 =================
        instruction_frame = ttk.LabelFrame(right_frame, text="使用说明", padding=10)
        instruction_frame.pack(fill=tk.BOTH, expand=True)

        # 更新使用说明文字内容
        instructions = """
        操作指南：

        1. 模板文件：选择元数据模板文件
        2. 边界数据：选择包含边界关系的Excel文件
        3. 目标数据：选择需要生成元数据的Excel文件
        4. TIFF文件夹：选择存放影像文件的目录
        5. 输出目录：指定元数据文件的生成位置
        
        功能说明：
        - 支持.xls和.xlsx格式输入，需保持格式一致
        - 支持文件拖拽到输入框
        - 自动校验输入文件格式
        - 生成日志实时查看

        注意事项：
        - 确保所有Excel文件符合格式要求
        - 输出目录建议使用空文件夹
        - 处理大文件时请耐心等待
        - 遇到错误请查看日志详情
        
        Copyright (c) 2025/07/25 Xiongxiao Wu.
        """

        # 带滚动条的文本框
        text_frame = ttk.Frame(instruction_frame)
        text_frame.pack(fill=tk.BOTH, expand=True)

        text = tk.Text(text_frame, wrap=tk.WORD, font=('微软雅黑', 9),
                       padx=5, pady=5, height=15)
        vsb = ttk.Scrollbar(text_frame, command=text.yview)
        text.configure(yscrollcommand=vsb.set)

        text.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")

        text.insert(tk.END, instructions)
        text.configure(state=tk.DISABLED)  # 设为只读模式

        # 配置文本框区域的网格权重
        text_frame.columnconfigure(0, weight=1)
        text_frame.rowconfigure(0, weight=1)

        # 状态栏
        self.status_var = tk.StringVar()
        status_bar = ttk.Label(self, textvariable=self.status_var, relief=tk.SUNKEN)
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)

    def _create_file_entry(self, parent, row, is_folder=False, command=None):
        """创建带拖放功能的文件输入组件"""
        frame = ttk.Frame(parent)
        frame.grid(row=row, column=1, sticky=tk.EW, pady=2)

        entry = ttk.Entry(frame)
        entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))

        btn_text = "选择目录" if is_folder else "选择文件"
        btn = ttk.Button(frame, text=btn_text, width=8,
                         command=command or (lambda: self.select_path(entry, is_folder)))
        btn.pack(side=tk.RIGHT)

        # 启用拖放功能
        entry.drop_target_register(DND_FILES)
        entry.dnd_bind('<<Drop>>', lambda e: self._handle_drop(e, entry, is_folder))
        return entry

    def _handle_drop(self, event, entry, is_folder):
        """处理文件拖放事件"""
        files = event.data.split()
        if files:
            path = files[0].strip('{}')  # 处理Windows路径的格式
            if is_folder:
                path = str(Path(path).parent) if Path(path).is_file() else path
            entry.delete(0, tk.END)
            entry.insert(0, path)

    def select_template(self):
        """选择模板文件"""
        path = filedialog.askopenfilename(
            title="选择模板文件",
            filetypes=[("Excel文件", "*.xlsx;*.xls"), ("所有文件", "*.*")]
        )
        if path:
            self.template_entry.delete(0, tk.END)
            self.template_entry.insert(0, path)

    def select_path(self, entry, is_folder):
        """通用路径选择方法"""
        initial = entry.get() or None
        if is_folder:
            path = filedialog.askdirectory(initialdir=initial)
        else:
            path = filedialog.askopenfilename(
                initialdir=initial,
                filetypes=[("Excel文件", "*.xlsx;*.xls"), ("所有文件", "*.*")]
            )
        if path:
            entry.delete(0, tk.END)
            entry.insert(0, path)

    def _setup_logging(self):
        """配置日志记录到文本框"""

        class TextHandler(logging.Handler):
            def __init__(self, text_widget):
                super().__init__()
                self.text_widget = text_widget

            def emit(self, record):
                msg = self.format(record)
                self.text_widget.configure(state=tk.NORMAL)
                self.text_widget.insert(tk.END, msg + '\n')
                self.text_widget.configure(state=tk.DISABLED)
                self.text_widget.see(tk.END)

        text_handler = TextHandler(self.log_text)
        text_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
        self.logger.addHandler(text_handler)
        self.logger.setLevel(logging.INFO)

    def validate_inputs(self):
        """验证输入参数"""
        required = {
            "模板文件": self.template_entry.get(),
            "边界数据文件": self.boundary_entry.get(),
            "目标数据文件": self.target_entry.get(),
            "TIFF文件夹": self.tiff_entry.get(),
            "输出目录": self.output_entry.get()
        }

        missing = [name for name, path in required.items() if not path.strip()]
        if missing:
            messagebox.showerror("输入错误", f"以下必填项未填写:\n{', '.join(missing)}")
            return False

        for name, path in required.items():
            if not Path(path).exists():
                messagebox.showerror("路径错误", f"{name}路径不存在: {path}")
                return False

        # 验证模板文件格式
        template_ext = Path(self.template_entry.get()).suffix.lower()
        if template_ext not in ['.xls', '.xlsx']:
            messagebox.showerror("格式错误", "模板文件必须是.xls或.xlsx格式")
            return False

        return True

    def start_processing(self):
        """启动处理流程"""
        if not self.validate_inputs():
            return

        try:
            self.status_var.set("正在初始化处理器...")
            self.update()

            self.processor = MetadataProcessor(self.template_entry.get())

            self.logger.info("正在加载数据...")
            self.processor.load_data(
                boundary_file=self.boundary_entry.get(),
                target_file=self.target_entry.get()
            )

            self.logger.info("正在生成元数据...")
            self.processor.generate_metadata(
                tif_folder=self.tiff_entry.get(),
                output_dir=self.output_entry.get()
            )

            self.status_var.set("处理完成")
            self.logger.info("元数据生成完成")
            messagebox.showinfo("完成", "元数据文件生成成功！")
            self.play_sound()

        except Exception as e:
            self.logger.error(f"处理失败: {str(e)}", exc_info=True)
            messagebox.showerror("处理错误", f"发生错误:\n{str(e)}")
            self.status_var.set("处理失败")
        finally:
            self.processor = None

    def play_sound(self):
        """任务完成提示音（Windows）"""
        try:
            import winsound
            winsound.MessageBeep()
        except:
            pass

    def reset_fields(self):
        """重置所有输入字段"""
        for entry in [self.template_entry, self.boundary_entry,
                      self.target_entry, self.tiff_entry, self.output_entry]:
            entry.delete(0, tk.END)
        self.log_text.configure(state=tk.NORMAL)
        self.log_text.delete(1.0, tk.END)
        self.log_text.configure(state=tk.DISABLED)
        self.status_var.set("就绪")


if __name__ == "__main__":
    app = MetadataProcessorGUI()
    app.mainloop()